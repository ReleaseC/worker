import os
import sys
import cv2
import numpy as np
import math
import glob
import argparse
from PIL import Image
from openpyxl import load_workbook

def parse_args():
	parser = argparse.ArgumentParser(description='Five points correction')

	parser.add_argument('-i', dest='coordinateXlsx', help='the xlsx file which describe center and top coordinates', default='./src/example')

	parser.add_argument('-s', dest='imgSrc', help='the directory of the source images', default='./src/example')

	parser.add_argument('-d', dest='imgDst', help='the directory of the final images', default='./output/example')

	parser.add_argument('-r', dest='replacement', help='the replacement order of the frame', default='')

	args = parser.parse_args()

	if not os.path.exists(args.imgSrc):
		print('data is not exist!')
		sys.exit()
	return args

args = parse_args()

# 1. Find center and top coordinates from xlsx
wb = load_workbook(args.coordinateXlsx)
sheets = wb.sheetnames
sheet0 = sheets[0]
ws = wb[sheet0]

row_count = ws.max_row

origin_center_x = []
origin_center_y = []
origin_top_x = []
origin_top_y = []

replacement = args.replacement.split(',')

for i in range(0, len(replacement)):
	# find center coordinate, 'XXX, YYY', split ', '
	tempCenter = ws.cell(row=int(replacement[i])+1, column=2).value.split(', ')
	origin_center_x.append(int(tempCenter[0]))
	origin_center_y.append(int(tempCenter[1]))
	# find top coordinate, 'XXX, YYY', split ', '
	tempTop = ws.cell(row=int(replacement[i])+1, column=3).value.split(', ')
	origin_top_x.append(int(tempTop[0]))
	origin_top_y.append(int(tempTop[1]))

# 2. Find number of source pics
listing = glob.glob(args.imgSrc+'/*.jpg')
n_images = len(listing)
# print 'number of images=',n_images
finalFrame = []

# 3. Start correction
for i in range(0, n_images):
	# print 'origin_center_x[%d]='%(i),origin_center_x[i]
	# print 'origin_center_xy%d]='%(i),origin_center_y[i]
	# print 'origin_top_x[%d]='%(i),origin_top_x[i]
    # print 'origin_top_y[%d]='%(i),origin_top_y[i]

	# Find distance between center and top
	dx = origin_center_x[i] - origin_top_x[i]
	dy = origin_center_y[i] - origin_top_y[i]
	dis = math.sqrt(dx*dx + dy*dy)

	# Load image
	canvas = cv2.imread(args.imgSrc+'/%d.jpg'%(i+1), cv2.IMREAD_COLOR)

	# Find center X, Y
	h, w, _ = canvas.shape
	cx = w / 2
	cy = h / 2

	# Find third point
	# x' = x*cos90 - y*sin90
	# y' = x*sin90 + y*cos90
	temp_p2_x = dx*0 - dy*1
	temp_p2_y = dx*1 + dy*0
	# Relative to center points
	origin_p2_x = origin_center_x[i] + temp_p2_x
	origin_p2_y = origin_center_y[i] + temp_p2_y

	# Start rotate and shift
	pts1 = np.float32([[origin_center_x[i],origin_center_y[i]],[origin_p2_x,origin_p2_y],[origin_top_x[i],origin_top_y[i]]])
	pts2 = np.float32([[cx,cy],[cx-dis,cy],[cx,cy-dis]])
	M = cv2.getAffineTransform(pts1,pts2)
	res = cv2.warpAffine(canvas,M,(w,h))

	# Write center point
	# cv2.circle(res, (cx, cy), 5, (0, 255, 0), 5)
	
	# Save to final directory
	im = Image.fromarray(res[:,:,::-1])
	with open(args.imgDst+'/%d_final.jpg'%(i+1), 'wb') as f:
		im.save(f, 'JPEG')
		finalFrame.append(args.imgDst+'/%d_final.jpg'%(i+1))
	f.closed
	with open(args.imgDst+'/%d_final.jpg'%(2*n_images-i), 'wb') as f:
		im.save(f, 'JPEG')
		finalFrame.append(args.imgDst+'/%d_final.jpg'%(2*n_images-i))
	f.closed


	# im.save(args.imgDst+'/%d_final.jpg'%(n_images*2-i))	
for i in range(0, 2*n_images):
	print(finalFrame[i])
